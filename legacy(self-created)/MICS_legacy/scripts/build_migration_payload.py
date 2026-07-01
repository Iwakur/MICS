from __future__ import annotations

import json
import math
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


KNOWN_FIRST_NAMES = {
    "андрій",
    "артем",
    "василь",
    "валерій",
    "віталій",
    "владислав",
    "гліб",
    "демо",
    "дмитро",
    "едуард",
    "євгеній",
    "ігор",
    "іраклі",
    "кіріл",
    "максим",
    "маргарита",
    "олександр",
    "олексій",
    "роман",
    "савенко",
    "тимур",
    "христина",
    "юрій",
}

ROLE_TRANSLATIONS = {
    "ютуб/тт продюсер": "YouTube/TikTok Producer",
    "бухгалтер": "Accountant",
    "асистент": "Assistant",
}

GROUP_PREFIX_RE = re.compile(r"^(D\d|A\d|M\d)\s+(.+)$", re.IGNORECASE)


@dataclass
class PlanDef:
    name: str
    lesson_count: float
    lesson_price: float
    teacher_share_per_lesson: float
    is_assignable: bool
    comments: str | None
    source: str


def normalize_space(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def normalize_key(value: str) -> str:
    return normalize_space(value).casefold()


def round2(value: Any) -> float:
    number = float(value or 0)
    return round(number + 1e-9, 2)


def round_lesson_count(value: Any) -> float:
    number = float(value or 0)
    return round(number + 1e-9, 2)


def plan_tuple(lesson_count: float, lesson_price: float, teacher_share_per_lesson: float) -> tuple[float, float, float]:
    return (round_lesson_count(lesson_count), round2(lesson_price), round2(teacher_share_per_lesson))


def approx(left: float | None, right: float | None, tolerance: float = 0.05) -> bool:
    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def split_formal_name(raw_value: Any) -> dict[str, str | None]:
    raw = normalize_space(raw_value)
    if raw == "":
        return {"family_name": None, "first_name": "", "father_name": None}

    tokens = raw.split()
    if len(tokens) == 1:
        return {"family_name": None, "first_name": tokens[0], "father_name": None}

    if len(tokens) == 2:
        left = tokens[0]
        right = tokens[1]
        if left.casefold() in KNOWN_FIRST_NAMES and right.casefold() not in KNOWN_FIRST_NAMES:
            return {"family_name": right, "first_name": left, "father_name": None}
        return {"family_name": left, "first_name": right, "father_name": None}

    return {
        "family_name": tokens[0],
        "first_name": tokens[1],
        "father_name": " ".join(tokens[2:]),
    }


def split_student_name(raw_value: Any) -> tuple[str | None, str, str | None, str | None]:
    raw = normalize_space(raw_value)
    if raw == "":
        return None, "", None, None

    prefix = None
    match = GROUP_PREFIX_RE.match(raw)
    if match:
        prefix = match.group(1).upper()
        raw = match.group(2).strip()

    if "/" in raw:
        parts = [normalize_space(part) for part in raw.split("/") if normalize_space(part) != ""]
        if parts:
            raw = max(parts, key=len)

    tokens = raw.split()
    if len(tokens) == 1:
        return prefix, tokens[0], None, None
    if len(tokens) == 2:
        left = tokens[0]
        right = tokens[1]
        if left.casefold() in KNOWN_FIRST_NAMES and right.casefold() not in KNOWN_FIRST_NAMES:
            return prefix, left, right, None
        return prefix, right, left, None

    if tokens[0].casefold() in KNOWN_FIRST_NAMES:
        return prefix, tokens[0], tokens[1], " ".join(tokens[2:])

    return prefix, tokens[1], tokens[0], " ".join(tokens[2:])


def build_full_name(parts: dict[str, str | None]) -> str:
    return " ".join(part for part in [parts.get("family_name"), parts.get("first_name"), parts.get("father_name")] if part)


def parse_staff_reference_rows(sheet) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], dict[str, str]]:
    workbook_rows: list[dict[str, Any]] = []
    exact_name_map: dict[str, dict[str, Any]] = {}
    first_name_map: dict[str, str] = {}

    if "stuff more info" not in sheet:
        return workbook_rows, exact_name_map, first_name_map

    ws = sheet["stuff more info"]

    for row in ws.iter_rows(values_only=True):
        text = normalize_space(row[0] if row else None)
        if text == "" or text.casefold().startswith("middle name"):
            continue

        match = re.match(r"^(.*?)(\d{16}|-+)$", text)
        if match:
            raw_name = normalize_space(match.group(1))
            raw_card = match.group(2)
            payout_card_number = raw_card if raw_card.strip("-") != "" else None
        else:
            raw_name = text
            payout_card_number = None

        parts = split_formal_name(raw_name)
        full_name = build_full_name(parts)
        entry = {
            "raw_name": raw_name,
            "full_name": full_name,
            "payout_card_number": payout_card_number,
            **parts,
        }
        workbook_rows.append(entry)

        if full_name != "":
            exact_name_map[normalize_key(full_name)] = entry
            first_name = normalize_key(parts["first_name"] or "")
            if first_name != "" and first_name not in first_name_map:
                first_name_map[first_name] = full_name

    return workbook_rows, exact_name_map, first_name_map


def resolve_staff_identity(
    raw_name: str,
    alias: str | None,
    exact_map: dict[str, dict[str, Any]],
    first_name_map: dict[str, str],
) -> tuple[dict[str, str | None], str | None, str]:
    parts = split_formal_name(raw_name)
    full_name = build_full_name(parts)
    exact = exact_map.get(normalize_key(full_name))
    if exact is not None:
        return (
            {
                "family_name": exact["family_name"],
                "first_name": exact["first_name"],
                "father_name": exact["father_name"],
            },
            exact["payout_card_number"],
            exact["full_name"],
        )

    if alias:
        alias_key = normalize_key(alias)
        matched_full_name = first_name_map.get(alias_key)
        if matched_full_name is not None:
            exact = exact_map.get(normalize_key(matched_full_name))
            if exact is not None:
                return (
                    {
                        "family_name": exact["family_name"],
                        "first_name": exact["first_name"],
                        "father_name": exact["father_name"],
                    },
                    exact["payout_card_number"],
                    exact["full_name"],
                )

    return parts, None, full_name


def add_staff(
    registry: OrderedDict[str, dict[str, Any]],
    *,
    family_name: str | None,
    first_name: str,
    father_name: str | None,
    role: str,
    payout_card_number: str | None,
    fixed_salary_amount: float | None,
    comments: str | None,
    alias: str | None = None,
) -> str:
    full_name = build_full_name(
        {
            "family_name": family_name,
            "first_name": first_name,
            "father_name": father_name,
        }
    )
    key = normalize_key(full_name if full_name != "" else first_name)
    if key in registry:
        current = registry[key]
        if current.get("payout_card_number") is None and payout_card_number is not None:
            current["payout_card_number"] = payout_card_number
        if current.get("fixed_salary_amount") is None and fixed_salary_amount is not None:
            current["fixed_salary_amount"] = fixed_salary_amount
        if alias and alias not in current["aliases"]:
            current["aliases"].append(alias)
        if comments and comments not in current["comment_parts"]:
            current["comment_parts"].append(comments)
        return key

    registry[key] = {
        "family_name": family_name,
        "first_name": first_name,
        "father_name": father_name,
        "role": role,
        "status": "active",
        "payout_card_number": payout_card_number,
        "fixed_salary_amount": fixed_salary_amount,
        "aliases": [alias] if alias else [],
        "comment_parts": [comments] if comments else [],
    }
    return key


def ensure_plan(
    plans: OrderedDict[str, PlanDef],
    name: str,
    lesson_count: float,
    lesson_price: float,
    teacher_share_per_lesson: float,
    is_assignable: bool,
    comments: str | None,
    source: str,
) -> str:
    candidate = normalize_space(name)
    if candidate == "":
        candidate = "Custom"

    suffix_index = 1
    while candidate in plans:
        existing = plans[candidate]
        if plan_tuple(existing.lesson_count, existing.lesson_price, existing.teacher_share_per_lesson) == plan_tuple(
            lesson_count,
            lesson_price,
            teacher_share_per_lesson,
        ):
            return candidate
        suffix_index += 1
        candidate = f"{name} #{suffix_index}"

    plans[candidate] = PlanDef(
        name=candidate,
        lesson_count=round_lesson_count(lesson_count),
        lesson_price=round2(lesson_price),
        teacher_share_per_lesson=round2(teacher_share_per_lesson),
        is_assignable=is_assignable,
        comments=comments,
        source=source,
    )
    return candidate


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")

    source_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\Hlib\Desktop\Книга1.xlsx")
    workbook_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(
        r"C:\laragon\www\GitHub\MICS\outputs\migration_workbook\MICS_Migration_Worksheet.xlsx"
    )
    output_path = Path(sys.argv[3]) if len(sys.argv) > 3 else Path(
        r"C:\laragon\www\GitHub\MICS\outputs\migration_workbook\migration_payload.json"
    )

    source_wb = load_workbook(source_path, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]

    workbook_wb = load_workbook(workbook_path, data_only=True)
    workbook_rows, workbook_exact_map, workbook_first_name_map = parse_staff_reference_rows(workbook_wb)

    plans: OrderedDict[str, PlanDef] = OrderedDict()

    header_row = [cell for cell in next(source_ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    price_row = [cell for cell in next(source_ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    lessons_row = [cell for cell in next(source_ws.iter_rows(min_row=3, max_row=3, values_only=True))]
    teacher_total_row = [cell for cell in next(source_ws.iter_rows(min_row=5, max_row=5, values_only=True))]

    canonical_by_name: dict[str, dict[str, float]] = {}
    for index in range(44, len(header_row)):
        raw_name = normalize_space(header_row[index])
        if raw_name == "":
            continue

        monthly_total = float(price_row[index] or 0)
        lesson_count = float(lessons_row[index] or 0)
        teacher_total = float(teacher_total_row[index] or 0)
        if lesson_count <= 0:
            continue

        lesson_price = round2(monthly_total / lesson_count)
        teacher_share_per_lesson = round2(teacher_total / lesson_count) if teacher_total > 0 else 0.0
        comments = f"Imported from tariff matrix. Monthly total {round2(monthly_total):.2f}; teacher monthly share {round2(teacher_total):.2f}."
        ensure_plan(
            plans,
            raw_name,
            lesson_count,
            lesson_price,
            teacher_share_per_lesson,
            True,
            comments,
            "matrix",
        )
        canonical_by_name[normalize_key(raw_name)] = {
            "lesson_count": round_lesson_count(lesson_count),
            "lesson_price": lesson_price,
            "teacher_share_per_lesson": teacher_share_per_lesson,
            "monthly_total": round2(monthly_total),
            "teacher_total": round2(teacher_total),
        }

    staff_registry: OrderedDict[str, dict[str, Any]] = OrderedDict()
    alias_to_staff_key: dict[str, str] = {}

    for row in source_ws.iter_rows(min_row=2, values_only=True):
        full_name = normalize_space(row[23])
        account_ref = normalize_space(row[24])
        raw_role = normalize_space(row[25])
        raw_salary = row[26]
        alias = normalize_space(row[24]) if normalize_space(row[24]) not in {"", "Див. BDB.txt"} else None
        teacher_count = row[27]

        if full_name == "" or full_name == "Підсумок" or full_name == "ПІБ":
            continue

        is_teacher_block = alias is not None or teacher_count is not None or raw_role == "Див. BDB.txt"

        if is_teacher_block:
            resolved_parts, payout_card_number, resolved_full_name = resolve_staff_identity(
                full_name,
                alias,
                workbook_exact_map,
                workbook_first_name_map,
            )
            key = add_staff(
                staff_registry,
                family_name=resolved_parts["family_name"],
                first_name=resolved_parts["first_name"] or alias or full_name,
                father_name=resolved_parts["father_name"],
                role="Teacher",
                payout_card_number=payout_card_number,
                fixed_salary_amount=None,
                comments="Imported from Excel teacher alias block. Teacher payouts are derived from plans.",
                alias=alias,
            )
            if alias:
                alias_to_staff_key[normalize_key(alias)] = key
            continue

        if raw_role == "":
            continue

        role = ROLE_TRANSLATIONS.get(raw_role.casefold(), raw_role)
        resolved_parts, payout_card_number, _ = resolve_staff_identity(
            full_name,
            None,
            workbook_exact_map,
            workbook_first_name_map,
        )
        fixed_salary_amount = round2(raw_salary) if raw_salary not in (None, "") else None
        add_staff(
            staff_registry,
            family_name=resolved_parts["family_name"],
            first_name=resolved_parts["first_name"] or full_name,
            father_name=resolved_parts["father_name"],
            role=role,
            payout_card_number=payout_card_number if payout_card_number is not None else (account_ref if re.fullmatch(r"\d{16}", account_ref) else None),
            fixed_salary_amount=fixed_salary_amount,
            comments="Imported from Excel fixed-salary block.",
        )

    for workbook_entry in workbook_rows:
        full_name = workbook_entry["full_name"] or workbook_entry["first_name"]
        key = normalize_key(full_name)
        if key in staff_registry:
            current = staff_registry[key]
            if current["payout_card_number"] is None and workbook_entry["payout_card_number"] is not None:
                current["payout_card_number"] = workbook_entry["payout_card_number"]
            continue

        add_staff(
            staff_registry,
            family_name=workbook_entry["family_name"],
            first_name=workbook_entry["first_name"] or full_name,
            father_name=workbook_entry["father_name"],
            role="Staff",
            payout_card_number=workbook_entry["payout_card_number"],
            fixed_salary_amount=None,
            comments="Imported from workbook only. Role was not explicitly specified.",
        )

    unassigned_teacher_key = add_staff(
        staff_registry,
        family_name="Owner",
        first_name="Unassigned",
        father_name=None,
        role="Teacher",
        payout_card_number=None,
        fixed_salary_amount=None,
        comments="Auto-created because some Excel rows did not specify a responsible teacher.",
        alias="UNASSIGNED",
    )
    alias_to_staff_key["__unassigned__"] = unassigned_teacher_key

    used_base_names: dict[str, tuple[float, float, float]] = {}
    custom_plan_index = 0
    students: list[dict[str, Any]] = []

    for row in source_ws.iter_rows(min_row=2, values_only=True):
        raw_id = row[0]
        raw_name = normalize_space(row[1])
        if raw_id is None or raw_name == "":
            continue

        prefix, first_name, family_name, father_name = split_student_name(raw_name)
        staff_alias = normalize_space(row[2])
        raw_plan = row[3]
        lesson_count_value = row[4]
        final_amount_value = row[5]
        discount_value = row[6]
        first_month_flag = bool(row[8])
        teacher_total_value = row[9]

        if isinstance(raw_id, (int, float)) and not isinstance(raw_id, bool):
            numeric_id = float(raw_id)
        else:
            numeric_id = 0.0

        if not math.isclose(numeric_id % 1, 0.0):
            status = "paused"
        elif numeric_id > 10:
            status = "archived"
        else:
            status = "active"

        lesson_count = round_lesson_count(lesson_count_value) if lesson_count_value not in (None, "") else 0.0
        final_amount = round2(final_amount_value) if final_amount_value not in (None, "") else 0.0
        discount_amount = round2(discount_value) if discount_value not in (None, "") else 0.0
        teacher_total = round2(teacher_total_value) if teacher_total_value not in (None, "") else 0.0

        raw_plan_label = normalize_space(raw_plan)
        raw_plan_number = float(raw_plan) if isinstance(raw_plan, (int, float)) and not isinstance(raw_plan, bool) else None
        plan_name: str

        canonical = canonical_by_name.get(normalize_key(raw_plan_label)) if raw_plan_label else None
        before_discount_total = round2(final_amount + discount_amount)

        if canonical and (
            lesson_count == 0.0 or approx(canonical["lesson_count"], lesson_count, 0.01)
        ) and (
            before_discount_total == 0.0 or approx(canonical["monthly_total"], before_discount_total)
        ) and (
            teacher_total == 0.0 or approx(canonical["teacher_total"], teacher_total)
        ):
            plan_name = raw_plan_label
        else:
            if lesson_count > 0:
                if raw_plan_number is not None and raw_plan_number > 0:
                    lesson_price = round2(raw_plan_number)
                    before_discount_total = round2(lesson_count * lesson_price)
                elif before_discount_total > 0:
                    lesson_price = round2(before_discount_total / lesson_count)
                else:
                    lesson_price = 0.0

                teacher_share_per_lesson = round2(teacher_total / lesson_count) if teacher_total > 0 else 0.0
            else:
                lesson_price = 0.0
                teacher_share_per_lesson = 0.0

            base_name = raw_plan_label if raw_plan_label else (f"{prefix} custom" if prefix else "Custom")
            if raw_plan_number is not None and raw_plan_label == "":
                base_name = f"{prefix} {int(raw_plan_number)}" if prefix else f"Custom {int(raw_plan_number)}"
            if lesson_price == 0.0 and teacher_share_per_lesson == 0.0 and before_discount_total == 0.0:
                base_name = f"{prefix} unresolved" if prefix else "Unresolved"

            candidate_name = normalize_space(base_name)
            tuple_value = plan_tuple(lesson_count, lesson_price, teacher_share_per_lesson)
            if candidate_name in used_base_names and used_base_names[candidate_name] != tuple_value:
                custom_plan_index += 1
                candidate_name = f"{candidate_name} [{lesson_count:g}/{before_discount_total:g}]"
                if candidate_name in plans and plan_tuple(
                    plans[candidate_name].lesson_count,
                    plans[candidate_name].lesson_price,
                    plans[candidate_name].teacher_share_per_lesson,
                ) != tuple_value:
                    candidate_name = f"{candidate_name} #{custom_plan_index}"
            else:
                used_base_names[candidate_name] = tuple_value

            plan_name = ensure_plan(
                plans,
                candidate_name,
                lesson_count,
                lesson_price,
                teacher_share_per_lesson,
                lesson_count > 0 and (lesson_price > 0 or before_discount_total > 0),
                (
                    f"Imported custom plan from student row {raw_id}. "
                    f"Source label={raw_plan_label or raw_plan_number or 'empty'}, "
                    f"monthly total before discount={before_discount_total:.2f}, "
                    f"teacher monthly share={teacher_total:.2f}."
                ),
                "student-custom",
            )

        if staff_alias != "":
            staff_key = alias_to_staff_key.get(normalize_key(staff_alias), unassigned_teacher_key)
        else:
            staff_key = unassigned_teacher_key

        students.append(
            {
                "raw_id": numeric_id,
                "raw_name": raw_name,
                "group_prefix": prefix,
                "family_name": family_name,
                "first_name": first_name,
                "father_name": father_name,
                "status": status,
                "plan_name": plan_name,
                "staff_key": staff_key,
                "discount_amount": discount_amount,
                "joined_at": "2026-04-01T00:00:00+03:00" if first_month_flag else "MIGRATION_TIMESTAMP",
            }
        )

    payload = {
        "metadata": {
            "source_workbook": str(source_path),
            "rules_workbook": str(workbook_path),
            "student_count": len(students),
            "staff_count": len(staff_registry),
            "plan_count": len(plans),
            "canonical_plan_count": len(canonical_by_name),
            "notes": [
                "April 2026 is the represented sheet month.",
                "Rows with decimal IDs map to paused status; IDs greater than 10 map to archived status.",
                "Teacher payouts remain derived from plans; non-teacher fixed salary amounts are stored on staff.",
                "Excel student comments and payment-confirmation flags are not imported.",
            ],
        },
        "staff": [
            {
                "key": key,
                "family_name": value["family_name"],
                "first_name": value["first_name"],
                "father_name": value["father_name"],
                "role": value["role"],
                "status": value["status"],
                "payout_card_number": value["payout_card_number"],
                "fixed_salary_amount": value["fixed_salary_amount"],
                "comments": " | ".join(value["comment_parts"]) if value["comment_parts"] else None,
                "aliases": value["aliases"],
            }
            for key, value in staff_registry.items()
        ],
        "plans": [
            {
                "name": plan.name,
                "lesson_count": plan.lesson_count,
                "lesson_price": plan.lesson_price,
                "teacher_share_per_lesson": plan.teacher_share_per_lesson,
                "is_assignable": plan.is_assignable,
                "comments": plan.comments,
                "source": plan.source,
            }
            for plan in plans.values()
        ],
        "students": students,
        "default_teacher_user_staff_key": next(
            (key for key, value in staff_registry.items() if value["role"].casefold() == "teacher"),
            unassigned_teacher_key,
        ),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
